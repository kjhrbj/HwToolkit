# -*- coding: utf-8 -*-
"""将 QueryoverSubcase 导出的 CSV 转成带高亮的 xlsx。

用法: write_excel_new-v2 <input.csv> <output.xlsx> [threshold]

可选 threshold 许用强度：
  提供时按 MaxValue / threshold 比值着色：
    >1 红 / 0.8~1 黄 / <0.8 不着色
  未提供时不进行着色。

与 write_excel_new.py 的区别：
  新增一个名为"总表"的 sheet，收集每个分表的第一行数据（即各组 MaxValue 最高的行），
  方便快速查看各组的最大值汇总。

分发说明：
  目标用户可能没有 Python，用 PyInstaller 打包成本目录 bin/ 下的 write_excel.exe：
    pip install openpyxl pyinstaller
    pyinstaller --onefile --name write_excel write_excel_new.py
  然后把 dist/write_excel.exe 放到 hv/scripts/bin/ 即可随插件分发。
"""
import csv
import re
import sys
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = "FFD9E1F2"


def color_for_threshold(v, threshold):
    """相对许用强度着色：<0.8 不着色 / 0.8~1 黄 / >1 红。"""
    ratio = v / threshold if threshold else 0.0
    if ratio > 1.0:
        return "FFFF0000"
    if ratio >= 0.8:
        return "FFFFFF00"
    return None


def sanitize_sheet_name(name, used_names):
    """将任意字符串转为合法的 Excel 分表名（<=31 字符，不含特殊字符，不重复）。"""
    name = str(name) if name else "(empty)"
    # 替换 Excel 禁止字符: [ ] : * ? / \
    name = re.sub(r'[\[\]:*?/\\]', '_', name)
    # 截断到 31 字符
    name = name[:31]
    # 去重：若重名则追加序号
    if name in used_names:
        base = name[:28]
        i = 2
        while f"{base}_{i}" in used_names:
            i += 1
        name = f"{base}_{i}"
    used_names.add(name)
    return name


def write_sheet(ws, header, rows, max_col, threshold):
    """将表头和数据行写入一个 sheet，含格式、着色、列宽自适应、冻结首行。"""
    # 表头：加粗 + 浅蓝底 + 居中
    for c, h in enumerate(header, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center")

    # 数据行 + 着色
    for ri, r in enumerate(rows, start=2):
        for c, val in enumerate(r, start=1):
            col_name = header[c - 1] if c - 1 < len(header) else ""
            if col_name == "MaxValue":
                try:
                    ws.cell(row=ri, column=c, value=int(round(float(val))))
                except (ValueError, TypeError):
                    ws.cell(row=ri, column=c, value=val)
            elif col_name == "@EntityID":
                try:
                    ws.cell(row=ri, column=c, value=int(float(val)))
                except (ValueError, TypeError):
                    ws.cell(row=ri, column=c, value=val)
            else:
                ws.cell(row=ri, column=c, value=val)

        # 只在使用许用强度时着色；留空或比值<0.8 则不填色
        if threshold is not None:
            try:
                v = float(r[max_col])
                fill = color_for_threshold(v, threshold)
                if fill:
                    ws.cell(row=ri, column=max_col + 1).fill = PatternFill(
                        "solid", fgColor=fill
                    )
            except (ValueError, IndexError):
                pass

    # 列宽自适应（最多看前 100 行）
    for c in range(1, len(header) + 1):
        width = 10
        for ri in range(1, min(len(rows) + 2, 101)):
            val = ws.cell(row=ri, column=c).value
            if val is not None:
                width = max(width, len(str(val)) + 2)
        ws.column_dimensions[get_column_letter(c)].width = width

    # 冻结首行
    ws.freeze_panes = "A2"


def main(csv_path, xlsx_path, threshold=None):
    if threshold:
        threshold = float(threshold)
    else:
        threshold = None

    with open(csv_path, "r", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))

    if not rows:
        raise SystemExit("CSV is empty")

    header = rows[0]
    # 找 MaxValue 所在列（找不到默认第 3 列）
    max_col = header.index("MaxValue") if "MaxValue" in header else 2

    # ---- 按第一列分组 ----
    groups = OrderedDict()
    for row in rows[1:]:
        key = row[0] if row and row[0] else "(empty)"
        groups.setdefault(key, []).append(row)

    # ---- 每组内按 MaxValue 降序排列 ----
    def sort_key(r):
        try:
            return float(r[max_col]) if max_col < len(r) and r[max_col] else float('-inf')
        except (ValueError, TypeError):
            return float('-inf')

    # 收集各组的首行（排序采用降序，首行即为 MaxValue 最大的行）
    summary_rows = []
    for group_rows in groups.values():
        group_rows.sort(key=sort_key, reverse=True)
        if group_rows:
            summary_rows.append(group_rows[0])

    # ---- 先创建总表（第一个 sheet） ----
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Overview"
    write_sheet(ws_summary, header, summary_rows, max_col, threshold)

    # ---- 为每组创建分表 ----
    used_names = set()
    used_names.add("总表")

    for key, group_rows in groups.items():
        sheet_name = sanitize_sheet_name(key, used_names)
        ws = wb.create_sheet(title=sheet_name)
        sub_header = header
        sub_rows = [r for r in group_rows]
        write_sheet(ws, sub_header, sub_rows, max_col, threshold)

    wb.save(xlsx_path)
    print(xlsx_path)


if __name__ == "__main__":
    if len(sys.argv) not in (3, 4):
        raise SystemExit("usage: write_excel_new-v2 <in.csv> <out.xlsx> [threshold]")
    threshold = sys.argv[3] if len(sys.argv) == 4 else None
    main(sys.argv[1], sys.argv[2], threshold)